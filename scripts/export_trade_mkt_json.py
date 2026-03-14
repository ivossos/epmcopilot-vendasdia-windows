#!/usr/bin/env python3
"""
Export Trade Marketing analysis from Excel to JSON for the dashboard.
Reads data/análise_trade_mkt.xlsx and writes data/trade_mkt.json.
"""

import json
import os

try:
    import openpyxl
except ImportError:
    print("Run: pip install openpyxl")
    raise SystemExit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(SCRIPT_DIR, '..', 'data')
INPUT_FILE = os.path.join(DATA_DIR, 'análise_trade_mkt.xlsx')
OUTPUT_FILE = os.path.join(DATA_DIR, 'trade_mkt.json')


def fmt(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v == int(v):
            return int(v)
        return v
    s = str(v).strip()
    return s if s else None


def export_base(ws):
    """Base sheet: filial-level KPIs (Venda, % Ating, Lucratividade, Margem)."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        return []
    # Row 0: metric names, Row 1: Realizado/Orçado, Row 2: Trabalho
    # Row 3+: data
    out = []
    for r in rows[3:]:
        filial = fmt(r[0])
        if not filial:
            continue
        venda_real = r[1]
        venda_orc = r[2]
        pct_ating_real = r[3]
        pct_ating_orc = r[4]
        lucro_real = r[5]
        lucro_orc = r[6]
        margem_real = r[7]
        margem_orc = r[8]
        out.append({
            'filial': filial,
            'venda_realizado': venda_real,
            'venda_orcado': venda_orc,
            'pct_ating_venda_real': pct_ating_real,
            'pct_ating_venda_orc': pct_ating_orc,
            'lucratividade_realizado': lucro_real,
            'lucratividade_orcado': lucro_orc,
            'margem_real': margem_real,
            'margem_orc': margem_orc,
        })
    return out


def export_setores(ws):
    """Setores sheet: sector hierarchy."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    out = []
    for r in rows[1:]:
        membro = fmt(r[0])
        pai = fmt(r[1])
        alias = fmt(r[6]) if len(r) > 6 else None
        if not membro:
            continue
        out.append({
            'membro': membro,
            'pai': pai,
            'alias': alias,
        })
    return out


def export_analise_setor_filial(ws):
    """Análise Setor x Filial: setor, mês, filial, Venda, Lucratividade."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return []
    # Row 1: year per column (col 4=2026, col 6=2025)
    anos = rows[1] if len(rows) > 1 else []
    ano_orc = anos[4] if len(anos) > 4 else None
    ano_real = anos[6] if len(anos) > 6 else None
    out = []
    for r in rows[4:]:
        setor = fmt(r[0])
        mes = fmt(r[1])
        filial = fmt(r[2])
        venda_orc = r[4]
        lucro_orc = r[5]
        venda_real = r[6]
        lucro_real = r[7]
        if not setor and not filial:
            continue
        out.append({
            'setor': setor or '',
            'mes': mes or '',
            'filial': filial or '',
            'ano_orcado': str(ano_orc) if ano_orc else None,
            'ano_realizado': str(ano_real) if ano_real else None,
            'venda_orcado': venda_orc,
            'lucratividade_orcado': lucro_orc,
            'venda_realizado': venda_real,
            'lucratividade_realizado': lucro_real,
        })
    return out


def export_analise_fornecedor(ws):
    """Análise Fornecedor: fornecedor, mês, Venda, Lucratividade."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return []
    anos = rows[1] if len(rows) > 1 else []
    ano_orc = anos[4] if len(anos) > 4 else None
    ano_real = anos[6] if len(anos) > 6 else None
    out = []
    for r in rows[4:]:
        fornecedor = fmt(r[0])
        mes = fmt(r[1])
        venda_orc = r[4]
        lucro_orc = r[5]
        venda_real = r[6]
        lucro_real = r[7]
        if not fornecedor:
            continue
        out.append({
            'fornecedor': fornecedor,
            'mes': mes or '',
            'ano_orcado': str(ano_orc) if ano_orc else None,
            'ano_realizado': str(ano_real) if ano_real else None,
            'venda_orcado': venda_orc,
            'lucratividade_orcado': lucro_orc,
            'venda_realizado': venda_real,
            'lucratividade_realizado': lucro_real,
        })
    return out


def main():
    if not os.path.exists(INPUT_FILE):
        print(f"Input not found: {INPUT_FILE}")
        raise SystemExit(1)

    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)

    data = {
        'base': export_base(wb['base']),
        'setores': export_setores(wb['setores']),
        'analise_setor_filial': export_analise_setor_filial(wb['análise_setor_filial']),
        'analise_fornecedor': export_analise_fornecedor(wb['análise_fornecedor']),
    }

    wb.close()

    os.makedirs(DATA_DIR, exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"Exported to {OUTPUT_FILE}")
    print(f"  base: {len(data['base'])} rows")
    print(f"  setores: {len(data['setores'])} rows")
    print(f"  analise_setor_filial: {len(data['analise_setor_filial'])} rows")
    print(f"  analise_fornecedor: {len(data['analise_fornecedor'])} rows")


if __name__ == '__main__':
    main()
