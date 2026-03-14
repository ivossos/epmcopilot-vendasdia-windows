#!/usr/bin/env python3
"""
Parse métricas para dashboard.xlsx → data/acompanhamento_diario.json
Sheet: "01.07.30 - Acompanhamento Diári"

Column mapping (rows 1-4 are multi-level headers):
  A : Filial
  B : Cota Mês        (Orçado)
  C : Cota            (Orçado)
  D : Venda           (Realizado)
  E : % Ating. Venda  (Orçado)
  F : Venda AA        (Realizado)
  G : % Var. Venda AA (Realizado)
  H : % Promoção      (Realizado)
  I : separator
  J : % Margem Objetiva (Orçado)
  K : % Margem Real     (Realizado)
  L : % Ating. Margem   (Orçado)
  M : Dif. Margem       (Orçado)
  N : Margem AA         (Realizado)
  O : Var. Margem AA    (Realizado)
  P : separator
  Q : Estoque           (Orçado)
  R : Cobertura Obj.    (Orçado)
  S : Cobertura Proj.   (Orçado)
  T : % Ating. Cobertura(Orçado)
  U : Dif. Cobertura    (Orçado)
  V : separator
  W : % Perda Obj.      (Orçado)
  X : % Perda Proj.     (Realizado)
  Y : % Ating. Perda    (Orçado)
  Z : Dif. Perda        (Orçado)
  AA: separator
  AB: Fluxo Atual       (Realizado)
  AC: Fluxo AA          (Realizado)
  AD: % Var. Fluxo      (Realizado)
  AE: Dif. Fluxo        (Realizado)
  AF: separator
  AG: Ticket Atual      (Realizado)
  AH: Ticket AA         (Realizado)
  AI: % Var. Ticket     (Realizado)
  AJ: Dif. Ticket       (Realizado)
"""

import json
import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE  = os.path.join(SCRIPT_DIR, '..', '..', 'Downloads',
                           'métricas para dashboard.xlsx')
OUTPUT_FILE = os.path.join(SCRIPT_DIR, '..', 'data', 'acompanhamento_diario.json')

# Fallback: look next to the script directory if above not found
if not os.path.exists(EXCEL_FILE):
    EXCEL_FILE = os.path.join(os.path.expanduser('~'), 'Downloads',
                              'métricas para dashboard.xlsx')

try:
    import openpyxl
except ImportError:
    print('❌  openpyxl not installed. Run: pip3 install openpyxl', file=sys.stderr)
    sys.exit(1)

def col_letter_to_idx(letter):
    """'A'→0, 'B'→1, ... 'AJ'→35"""
    result = 0
    for c in letter.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result - 1

# Column → field name mapping
COL_MAP = {
    'A': 'filial',
    'B': 'cota_mes',
    'C': 'cota',
    'D': 'venda',
    'E': 'pct_ating_venda',
    'F': 'venda_aa',
    'G': 'pct_var_venda_aa',
    'H': 'pct_promocao',
    'J': 'margem_obj',
    'K': 'margem_real',
    'L': 'pct_ating_margem',
    'M': 'dif_margem',
    'N': 'margem_aa',
    'O': 'var_margem_aa',
    'Q': 'estoque',
    'R': 'cobertura_obj',
    'S': 'cobertura_proj',
    'T': 'pct_ating_cobertura',
    'U': 'dif_cobertura',
    'W': 'pct_perda_obj',
    'X': 'pct_perda_proj',
    'Y': 'pct_ating_perda',
    'Z': 'dif_perda',
    'AB': 'fluxo_atual',
    'AC': 'fluxo_aa',
    'AD': 'pct_var_fluxo',
    'AE': 'dif_fluxo',
    'AG': 'ticket_atual',
    'AH': 'ticket_aa',
    'AI': 'pct_var_ticket',
    'AJ': 'dif_ticket',
}

def safe_float(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def safe_str(v):
    if v is None:
        return ''
    return str(v).strip()

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f'❌  Excel não encontrado: {EXCEL_FILE}', file=sys.stderr)
        print(f'    Coloque "métricas para dashboard.xlsx" em ~/Downloads/', file=sys.stderr)
        sys.exit(1)

    print(f'📂  Lendo {EXCEL_FILE}')
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Find the sheet
    target_sheet = None
    for name in wb.sheetnames:
        if '01.07' in name or 'Acompanha' in name:
            target_sheet = name
            break
    if not target_sheet:
        # Fall back to second sheet
        target_sheet = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]

    ws = wb[target_sheet]
    print(f'📋  Sheet: {target_sheet}  ({ws.max_row} linhas × {ws.max_column} colunas)')

    # Build col_index → field map
    col_idx_map = {}
    for col_letter, field in COL_MAP.items():
        idx = col_letter_to_idx(col_letter) + 1  # openpyxl is 1-based
        col_idx_map[idx] = field

    # Data starts at row 5 (rows 1-4 are headers)
    DATA_START = 5
    rows = []
    totals = {}

    for row in ws.iter_rows(min_row=DATA_START, values_only=True):
        filial_val = safe_str(row[0]) if row else ''
        if not filial_val:
            continue

        rec = {}
        for col_idx_1based, field in col_idx_map.items():
            raw = row[col_idx_1based - 1] if col_idx_1based - 1 < len(row) else None
            if field == 'filial':
                rec[field] = safe_str(raw)
            else:
                rec[field] = safe_float(raw)

        # Classify row type
        fil = rec['filial']
        if fil.startswith('Totalizador') or fil.startswith('Total') or fil.upper().startswith('TOTAL'):
            totals[fil] = rec
        else:
            rows.append(rec)

    # Metadata: derive reference period from sheet name or hardcode
    out = {
        'reference': {
            'month': 'Fevereiro',
            'year': '2026',
            'scenario': 'Real/Trabalho',
            'filial': '01 - SVG',
            'canal': 'Total Canal',
            'categoria': 'Total Categoria',
        },
        'rows': rows,
        'totals': totals,
        'generated_at': __import__('datetime').datetime.now().isoformat(timespec='seconds'),
    }

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f'✅  {len(rows)} filiais + {len(totals)} totais → {OUTPUT_FILE}')

if __name__ == '__main__':
    main()
